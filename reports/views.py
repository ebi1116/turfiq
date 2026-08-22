from collections import defaultdict
from datetime import datetime
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.http import Http404
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle

from dashboard.services import date_bounds, report_data

HEADER_FILL = PatternFill("solid", fgColor="128A4B")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_BORDER = Border(bottom=Side(style="medium", color="0B5D32"))
BODY_BORDER = Border(bottom=Side(style="thin", color="DDE8E1"))
ALTERNATE_FILL = PatternFill("solid", fgColor="F0F8F3")
TITLE_FONT = Font(name="Calibri", size=20, bold=True, color="0B5D32")
MONEY_FORMAT = "\u20b9#,##0.00"


def _style_data_sheet(sheet, widths, money_columns=(), table_name=None):
    for cell in sheet[1]:
        cell.fill, cell.font, cell.border = HEADER_FILL, HEADER_FONT, HEADER_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 26
    for row_number in range(2, sheet.max_row + 1):
        for cell in sheet[row_number]:
            cell.font = Font(name="Calibri", size=11, color="24352B")
            cell.border = BODY_BORDER
            cell.alignment = Alignment(vertical="center")
            if row_number % 2 == 0:
                cell.fill = ALTERNATE_FILL
        sheet.row_dimensions[row_number].height = 21
    for column in money_columns:
        for column_cells in sheet.iter_cols(
            min_col=column, max_col=column, min_row=2, max_row=sheet.max_row
        ):
            for cell in column_cells:
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center")
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    sheet.print_title_rows = "1:1"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth, sheet.page_setup.fitToHeight = 1, 0
    sheet.sheet_view.zoomScale = 90
    if table_name and sheet.max_row > 1:
        table = ExcelTable(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        sheet.add_table(table)


def _write_pivot(sheet, start_row, title, headers, rows, money_columns=()):
    sheet.cell(start_row, 1, title).font = Font(name="Calibri", size=14, bold=True, color="0B5D32")
    header_row = start_row + 1
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(header_row, column, value)
        cell.fill, cell.font, cell.border = HEADER_FILL, HEADER_FONT, HEADER_BORDER
    for row_number, values in enumerate(rows, header_row + 1):
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_number, column, value)
            cell.border = BODY_BORDER
            if row_number % 2 == 0:
                cell.fill = ALTERNATE_FILL
            if column in money_columns:
                cell.number_format = MONEY_FORMAT
    return header_row, header_row + len(rows)


def _add_chart(sheet, chart, header_row, last_row, value_col, category_col, title, position):
    if last_row <= header_row:
        return
    chart.add_data(Reference(sheet, min_col=value_col, min_row=header_row, max_row=last_row), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=category_col, min_row=header_row + 1, max_row=last_row))
    chart.title, chart.height, chart.width, chart.legend = title, 7.5, 13, None
    sheet.add_chart(chart, position)


def _build_excel_report(data, start, end):
    bookings, expenses = list(data["bookings"]), list(data["expenses"])
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    dashboard.sheet_view.showGridLines = False
    dashboard.merge_cells("A1:H2")
    dashboard["A1"] = "TurfIQ Business Report"
    dashboard["A1"].font, dashboard["A1"].alignment = TITLE_FONT, Alignment(vertical="center")
    dashboard["A3"] = f"Reporting period: {start:%d %b %Y} to {end:%d %b %Y}"
    dashboard["A3"].font = Font(italic=True, color="607067")
    revenue_bookings = [booking for booking in bookings if booking.status != "Cancelled"]
    pending = sum((booking.amount for booking in revenue_bookings if not booking.is_paid), 0)
    kpis = [
        ("Revenue", float(data["revenue"]), True), ("Expenses", float(data["expense_total"]), True),
        ("Net Profit", float(data["profit"]), True), ("Bookings", len(bookings), False),
        ("Pending Collection", float(pending), True),
        ("Avg. Booking Value", float(data["revenue"] / max(len(revenue_bookings), 1)), True),
    ]
    for index, (label, value, is_money) in enumerate(kpis):
        row, column = 5 + (index // 3) * 3, 1 + (index % 3) * 3
        dashboard.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
        dashboard.merge_cells(start_row=row + 1, start_column=column, end_row=row + 1, end_column=column + 1)
        dashboard.cell(row, column, label).font = Font(bold=True, color="607067")
        dashboard.cell(row + 1, column, value).font = Font(size=16, bold=True, color="0B5D32")
        dashboard.cell(row, column).fill = dashboard.cell(row + 1, column).fill = PatternFill("solid", fgColor="EAF5EE")
        if is_money:
            dashboard.cell(row + 1, column).number_format = MONEY_FORMAT
    dashboard.merge_cells("A13:H13")
    dashboard["A13"] = "Use Pivot Tables for analysis and table filters in Bookings / Expenses for drill-down."
    dashboard["A13"].font = Font(italic=True, color="607067")
    dashboard.freeze_panes = "A4"
    for column in range(1, 9):
        dashboard.column_dimensions[get_column_letter(column)].width = 16

    booking_sheet = workbook.create_sheet("Bookings")
    booking_sheet.append(["Date", "Time", "Customer", "Phone", "Sport", "Ground", "Duration (hrs)", "Amount", "Payment", "Paid", "Status", "Notes"])
    for booking in bookings:
        booking_sheet.append([booking.booking_date, booking.booking_time, booking.customer.name, booking.customer.phone,
            booking.sport, str(booking.ground), float(booking.duration), float(booking.amount), booking.payment_method,
            "Yes" if booking.is_paid else "No", booking.status, booking.notes])
    _style_data_sheet(booking_sheet, [14, 12, 24, 17, 18, 22, 16, 16, 16, 10, 16, 32], (8,), "BookingsTable")
    for cell in booking_sheet["A"][1:]: cell.number_format = "dd-mmm-yyyy"
    for cell in booking_sheet["B"][1:]: cell.number_format = "hh:mm AM/PM"

    expense_sheet = workbook.create_sheet("Expenses")
    expense_sheet.append(["Date", "Category", "Amount", "Notes"])
    for expense in expenses:
        expense_sheet.append([expense.expense_date, expense.category, float(expense.amount), expense.notes])
    _style_data_sheet(expense_sheet, [14, 30, 16, 42], (3,), "ExpensesTable")
    for cell in expense_sheet["A"][1:]: cell.number_format = "dd-mmm-yyyy"

    by_month = defaultdict(lambda: {"bookings": 0, "revenue": 0.0, "expenses": 0.0})
    by_ground = defaultdict(lambda: {"bookings": 0, "revenue": 0.0})
    by_payment = defaultdict(lambda: {"bookings": 0, "revenue": 0.0})
    by_status = defaultdict(lambda: {"bookings": 0, "amount": 0.0})
    by_category = defaultdict(float)
    for booking in bookings:
        month = booking.booking_date.strftime("%b %Y")
        by_month[month]["bookings"] += 1
        by_status[booking.status]["bookings"] += 1
        by_status[booking.status]["amount"] += float(booking.amount)
        if booking.status != "Cancelled":
            by_month[month]["revenue"] += float(booking.amount)
            by_ground[str(booking.ground)]["bookings"] += 1
            by_ground[str(booking.ground)]["revenue"] += float(booking.amount)
            by_payment[booking.payment_method]["bookings"] += 1
            by_payment[booking.payment_method]["revenue"] += float(booking.amount)
    for expense in expenses:
        by_month[expense.expense_date.strftime("%b %Y")]["expenses"] += float(expense.amount)
        by_category[expense.category] += float(expense.amount)

    pivot = workbook.create_sheet("Pivot Tables")
    pivot.sheet_view.showGridLines = False
    month_keys = sorted(by_month, key=lambda value: datetime.strptime(value, "%b %Y"))
    month_rows = [[key, by_month[key]["bookings"], by_month[key]["revenue"], by_month[key]["expenses"], by_month[key]["revenue"] - by_month[key]["expenses"]] for key in month_keys]
    month_header, month_end = _write_pivot(pivot, 1, "Monthly Performance", ["Month", "Bookings", "Revenue", "Expenses", "Profit"], month_rows, (3, 4, 5))
    ground_start = month_end + 3
    _write_pivot(pivot, ground_start, "Ground Performance", ["Ground", "Bookings", "Revenue"], [[key, value["bookings"], value["revenue"]] for key, value in sorted(by_ground.items())], (3,))
    payment_start = ground_start + len(by_ground) + 4
    payment_header, payment_end = _write_pivot(pivot, payment_start, "Payment Method", ["Payment", "Bookings", "Revenue"], [[key, value["bookings"], value["revenue"]] for key, value in sorted(by_payment.items())], (3,))
    status_start = payment_end + 3
    _write_pivot(pivot, status_start, "Booking Status", ["Status", "Bookings", "Gross Amount"], [[key, value["bookings"], value["amount"]] for key, value in sorted(by_status.items())], (3,))
    category_start = status_start + len(by_status) + 4
    category_header, category_end = _write_pivot(pivot, category_start, "Expense Categories", ["Category", "Expense"], [[key, value] for key, value in sorted(by_category.items(), key=lambda item: item[1], reverse=True)], (2,))
    pivot.column_dimensions["A"].width = 30
    for column in "BCDE": pivot.column_dimensions[column].width = 16
    pivot.freeze_panes = "A3"
    _add_chart(pivot, LineChart(), month_header, month_end, 3, 1, "Monthly Revenue Trend", "G2")
    _add_chart(pivot, DoughnutChart(), payment_header, payment_end, 2, 1, "Bookings by Payment Method", "G18")
    _add_chart(pivot, BarChart(), category_header, category_end, 2, 1, "Expenses by Category", "G34")
    return workbook


def _pdf_table(rows, widths=None, money_columns=()):
    table = PdfTable(rows, colWidths=widths, repeatRows=1, hAlign="LEFT")
    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#128A4B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
        ("TOPPADDING", (0, 0), (-1, 0), 7),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#DDE8E1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F8F3")]),
    ]
    for column in money_columns:
        commands.append(("ALIGN", (column, 1), (column, -1), "RIGHT"))
    table.setStyle(TableStyle(commands))
    return table


def _pdf_page(canvas, document):
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor("#DDE8E1"))
    canvas.line(document.leftMargin, 12 * mm, A4[1] - document.rightMargin, 12 * mm)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#607067"))
    canvas.drawString(document.leftMargin, 8 * mm, "TurfIQ Business Report")
    canvas.drawRightString(A4[1] - document.rightMargin, 8 * mm, f"Page {document.page}")
    canvas.restoreState()


def _build_pdf_report(data, start, end):
    bookings, expenses = list(data["bookings"]), list(data["expenses"])
    output = BytesIO()
    document = SimpleDocTemplate(
        output, pagesize=landscape(A4), rightMargin=14 * mm, leftMargin=14 * mm,
        topMargin=14 * mm, bottomMargin=18 * mm,
        title="TurfIQ Business Report", author="TurfIQ",
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle("ReportTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=22, leading=26, textColor=colors.HexColor("#0B5D32"), alignment=TA_CENTER, spaceAfter=5 * mm)
    heading = ParagraphStyle("ReportHeading", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=14, leading=18, textColor=colors.HexColor("#0B5D32"), spaceBefore=5 * mm, spaceAfter=3 * mm)
    note = ParagraphStyle("ReportNote", parent=styles["BodyText"], fontSize=9, textColor=colors.HexColor("#607067"), alignment=TA_CENTER, spaceAfter=5 * mm)
    money = lambda value: f"INR {float(value):,.2f}"
    revenue_bookings = [booking for booking in bookings if booking.status != "Cancelled"]
    pending = sum((booking.amount for booking in revenue_bookings if not booking.is_paid), 0)
    average = data["revenue"] / max(len(revenue_bookings), 1)
    story = [
        Paragraph("TurfIQ Business Report", title),
        Paragraph(f"Reporting period: {start:%d %b %Y} to {end:%d %b %Y}", note),
        _pdf_table([
            ["Revenue", "Expenses", "Net Profit", "Bookings", "Pending Collection", "Avg. Booking Value"],
            [money(data["revenue"]), money(data["expense_total"]), money(data["profit"]), str(len(bookings)), money(pending), money(average)],
        ], [42 * mm, 42 * mm, 42 * mm, 30 * mm, 46 * mm, 46 * mm]),
        Paragraph("Booking Ledger", heading),
    ]
    booking_rows = [["Date", "Time", "Customer", "Sport", "Ground", "Duration", "Amount", "Payment", "Paid", "Status"]]
    booking_rows.extend([
        [booking.booking_date.strftime("%d-%b-%Y"), booking.booking_time.strftime("%I:%M %p"), booking.customer.name[:28], booking.sport,
         str(booking.ground)[:22], f"{booking.duration} hrs", money(booking.amount), booking.payment_method,
         "Yes" if booking.is_paid else "No", booking.status]
        for booking in bookings
    ])
    if len(booking_rows) == 1:
        booking_rows.append(["No bookings in the selected period", "", "", "", "", "", "", "", "", ""])
    story.extend([_pdf_table(booking_rows, [23*mm, 20*mm, 39*mm, 25*mm, 32*mm, 20*mm, 28*mm, 24*mm, 15*mm, 24*mm], (6,)), PageBreak(), Paragraph("Expense Ledger", heading)])
    expense_rows = [["Date", "Category", "Amount", "Notes"]]
    expense_rows.extend([[expense.expense_date.strftime("%d-%b-%Y"), expense.category, money(expense.amount), expense.notes[:90]] for expense in expenses])
    if len(expense_rows) == 1:
        expense_rows.append(["No expenses in the selected period", "", "", ""])
    story.append(_pdf_table(expense_rows, [32*mm, 58*mm, 38*mm, 138*mm], (2,)))

    by_ground, by_payment, by_status, by_category = defaultdict(lambda: [0, 0.0]), defaultdict(lambda: [0, 0.0]), defaultdict(lambda: [0, 0.0]), defaultdict(float)
    for booking in bookings:
        by_status[booking.status][0] += 1
        by_status[booking.status][1] += float(booking.amount)
        if booking.status != "Cancelled":
            by_ground[str(booking.ground)][0] += 1; by_ground[str(booking.ground)][1] += float(booking.amount)
            by_payment[booking.payment_method][0] += 1; by_payment[booking.payment_method][1] += float(booking.amount)
    for expense in expenses:
        by_category[expense.category] += float(expense.amount)
    analyses = [
        ("Ground Performance", [["Ground", "Bookings", "Revenue"]] + [[key, value[0], money(value[1])] for key, value in sorted(by_ground.items())], [65*mm, 35*mm, 50*mm], (2,)),
        ("Payment Methods", [["Payment", "Bookings", "Revenue"]] + [[key, value[0], money(value[1])] for key, value in sorted(by_payment.items())], [65*mm, 35*mm, 50*mm], (2,)),
        ("Booking Status", [["Status", "Bookings", "Gross Amount"]] + [[key, value[0], money(value[1])] for key, value in sorted(by_status.items())], [65*mm, 35*mm, 50*mm], (2,)),
        ("Expense Categories", [["Category", "Expense"]] + [[key, money(value)] for key, value in sorted(by_category.items(), key=lambda item: item[1], reverse=True)], [90*mm, 60*mm], (1,)),
    ]
    story.extend([PageBreak(), Paragraph("Business Analysis", heading)])
    for section_title, rows, widths, money_columns in analyses:
        if len(rows) == 1:
            rows.append(["No data", *([""] * (len(rows[0]) - 1))])
        story.append(KeepTogether([Paragraph(section_title, heading), _pdf_table(rows, widths, money_columns), Spacer(1, 3 * mm)]))
    document.build(story, onFirstPage=_pdf_page, onLaterPages=_pdf_page)
    return output.getvalue()


@login_required
def report_view(request):
    start, end, preset = date_bounds(request)
    data = report_data(request.user, start, end)
    data["preset"] = preset
    return render(request, "reports/index.html", data)


@login_required
def export_report(request, format):
    start, end, _ = date_bounds(request)
    data = report_data(request.user, start, end)
    rows = data["bookings"]
    if format == "xlsx":
        output = BytesIO()
        _build_excel_report(data, start, end).save(output)
        response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="turfiq-report.xlsx"'
        return response
    if format == "pdf":
        response = HttpResponse(_build_pdf_report(data, start, end), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="turfiq-report.pdf"'
        return response
    raise Http404("Unsupported report format")
